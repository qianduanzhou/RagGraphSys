"""services.file_parser 的测试。

文本 / 代码类用真实字节；PDF 用 monkeypatch 注入假 pypdf；
Word(.docx) 用真实 python-docx 做往返（未装则跳过）。
"""
import io
import types

import pytest

from services.file_parser import ALLOWED_EXTS, parse_upload, parse_upload_detail


# ------------------------------------------------------------------ #
# 文本 / 代码 / 配置类
# ------------------------------------------------------------------ #
@pytest.mark.parametrize("name", ["a.txt", "main.py", "App.tsx", "Main.java",
                                  "index.html", "style.css", "conf.json", "notes.md"])
def test_parse_text_code(name):
    assert parse_upload(name, "print('hello')".encode()) == "print('hello')"


def test_parse_text_strips_whitespace():
    assert parse_upload("a.py", b"   code here\n\n") == "code here"


def test_parse_code_handles_non_utf8_bytes():
    # 非 UTF-8 字节被 ignore，不抛错
    out = parse_upload("a.py", b"ok \xff\xfe end")
    assert "ok" in out and "end" in out


# ------------------------------------------------------------------ #
# PDF（monkeypatch pypdf，避免依赖真实 PDF 生成库）
# ------------------------------------------------------------------ #
def test_parse_pdf(monkeypatch):
    fake_page = types.SimpleNamespace(extract_text=lambda: "Page one text")
    fake_reader = lambda stream: types.SimpleNamespace(pages=[fake_page])
    fake_mod = types.ModuleType("pypdf")
    fake_mod.PdfReader = fake_reader
    monkeypatch.setitem(__import__("sys").modules, "pypdf", fake_mod)

    assert parse_upload("doc.pdf", b"%PDF- fake") == "Page one text"
    assert ".pdf" in ALLOWED_EXTS


def test_parse_pdf_expands_merged_table_notes(monkeypatch):
    note = "\u9002\u5e94\u9ad8\u7a7a\u4f5c\u4e1a\u3002"
    header = "\u5c0f\u8ba1 30"
    footer = "\u5c0f\u8ba1 80"
    text = (
        f"{header} "
        "62 G2026020701 Guangzhou power supply maintenance Guangdong 34 "
        f"{note} "
        "63 G2026020702 Shenzhen power supply maintenance Guangdong 10 "
        "64 G2026020703 Huizhou power supply maintenance Guangdong 4 "
        f"{footer}"
    )
    fake_page = types.SimpleNamespace(extract_text=lambda: text)
    fake_reader = lambda stream: types.SimpleNamespace(pages=[fake_page])
    fake_mod = types.ModuleType("pypdf")
    fake_mod.PdfReader = fake_reader
    monkeypatch.setitem(__import__("sys").modules, "pypdf", fake_mod)

    parsed = parse_upload("jobs.pdf", b"%PDF- fake")

    assert "\u3010PDF\u8868\u683c\u5408\u5e76\u5907\u6ce8\u5c55\u5f00\u3011" in parsed
    assert parsed.count(f"\u5907\u6ce8\uff1a{note}") == 3
    assert "G2026020702" in parsed
    assert "G2026020703" in parsed


def test_parse_pdf_expands_merged_table_notes_for_vision_requirements(monkeypatch):
    note = "\u65e0\u9ad8\u5ea6\u8fd1\u89c6\uff0c\u8eab\u4f53\u7d20\u8d28\n\u597d\u3002"
    normalized_note = "\u65e0\u9ad8\u5ea6\u8fd1\u89c6\uff0c\u8eab\u4f53\u7d20\u8d28 \u597d\u3002"
    header = "\u5c0f\u8ba1 52"
    footer = "\u5c0f\u8ba1 96"
    text = (
        f"{header} "
        "35 G2026020401 Guangzhou locomotive driving maintenance Guangdong 16 "
        "mechanical engineering automation "
        f"{note} "
        "36 G2026020402 Guangzhou locomotive overhaul Guangdong 9 "
        "37 G2026020403 Shantou locomotive driving Guangdong 12 "
        f"{footer}"
    )
    fake_page = types.SimpleNamespace(extract_text=lambda: text)
    fake_reader = lambda stream: types.SimpleNamespace(pages=[fake_page])
    fake_mod = types.ModuleType("pypdf")
    fake_mod.PdfReader = fake_reader
    monkeypatch.setitem(__import__("sys").modules, "pypdf", fake_mod)

    parsed = parse_upload("jobs.pdf", b"%PDF- fake")

    assert parsed.count(f"\u5907\u6ce8\uff1a{normalized_note}") == 3
    assert "G2026020402" in parsed
    assert "G2026020403" in parsed


def test_parse_pdf_adds_structured_job_records(monkeypatch):
    note = "\u65e0\u9ad8\u5ea6\u8fd1\u89c6\uff0c\u8eab\u4f53\u7d20\u8d28\u597d\u3002"
    text = (
        "\u5c0f\u8ba1 52 "
        "35 G2026020401 Guangzhou locomotive driving maintenance Guangdong 16 "
        f"{note} "
        "36 G2026020402 Guangzhou locomotive overhaul Guangdong 9 "
        "\u5c0f\u8ba1 96"
    )
    fake_page = types.SimpleNamespace(extract_text=lambda: text)
    fake_reader = lambda stream: types.SimpleNamespace(pages=[fake_page])
    fake_mod = types.ModuleType("pypdf")
    fake_mod.PdfReader = fake_reader
    monkeypatch.setitem(__import__("sys").modules, "pypdf", fake_mod)

    parsed = parse_upload("jobs.pdf", b"%PDF- fake")

    assert "\u3010PDF\u5c97\u4f4d\u884c\u7ed3\u6784\u5316\u7d22\u5f15\u3011" in parsed
    assert "\u5c97\u4f4d\u7f16\u53f7=G2026020402" in parsed
    assert f"\u5907\u6ce8={note}" in parsed


def test_parse_pdf_missing_lib(monkeypatch):
    # 未装 pypdf 时给出可读错误
    import sys
    monkeypatch.setitem(sys.modules, "pypdf", None)
    with pytest.raises(ValueError, match="pypdf"):
        parse_upload("doc.pdf", b"%PDF-")


# ------------------------------------------------------------------ #
# Word(.docx) —— 真实往返
# ------------------------------------------------------------------ #
def test_parse_docx_roundtrip():
    docx = pytest.importorskip("docx")
    doc = docx.Document()
    doc.add_paragraph("Hello from Word")
    buf = io.BytesIO()
    doc.save(buf)
    text = parse_upload("note.docx", buf.getvalue())
    assert "Hello from Word" in text
    assert ".docx" in ALLOWED_EXTS


def test_parse_docx_includes_header_and_footer():
    docx = pytest.importorskip("docx")
    doc = docx.Document()
    section = doc.sections[0]
    section.header.paragraphs[0].text = "Confidential Header"
    section.footer.paragraphs[0].text = "Footer Page"
    doc.add_paragraph("Body paragraph")
    buf = io.BytesIO()
    doc.save(buf)

    text = parse_upload("rich.docx", buf.getvalue())

    assert "Body paragraph" in text
    assert "Confidential Header" in text
    assert "Footer Page" in text


# ------------------------------------------------------------------ #
# CSV —— 结构化解析为 markdown 表格（保留表头与行列结构）
# ------------------------------------------------------------------ #
def test_parse_csv_renders_markdown_table():
    raw = b"name,age\nAlice,30\nBob,25\n"
    text = parse_upload("data.csv", raw)
    assert text.splitlines()[0] == "| name | age |"
    assert "| --- | --- |" in text
    assert "| Alice | 30 |" in text
    assert "| Bob | 25 |" in text


def test_parse_csv_adds_structured_records():
    text = parse_upload("data.csv", b"name,age\nAlice,30\nBob,25\n")

    assert "\u3010\u7ed3\u6784\u5316\u8868\u683c\u8bb0\u5f55\uff1aCSV\u3011" in text
    assert "\u8bb0\u5f551: name=Alice | age=30" in text
    assert "\u8bb0\u5f552: name=Bob | age=25" in text


def test_parse_csv_is_structured_not_raw_text():
    # csv 不应再是原始逗号串，应含表格分隔符
    assert "|" in parse_upload("data.csv", b"a,b\n1,2\n")


def test_csv_in_allowed_exts():
    assert ".csv" in ALLOWED_EXTS


def test_parse_csv_handles_gb18030_bytes():
    raw = "姓名,城市\n张三,广州\n".encode("gb18030")
    text = parse_upload("cn.csv", raw)
    assert "张三" in text
    assert "广州" in text


# ------------------------------------------------------------------ #
# Excel(.xlsx) —— openpyxl 往返，渲染为每 sheet 一个 markdown 表格
# ------------------------------------------------------------------ #
def test_parse_xlsx_renders_sheet_as_table():
    openpyxl = pytest.importorskip("openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "销售"
    ws.append(["产品", "数量"])
    ws.append(["苹果", 10])
    ws.append(["香蕉", 5])
    buf = io.BytesIO()
    wb.save(buf)
    text = parse_upload("data.xlsx", buf.getvalue())
    assert "## Sheet: 销售" in text
    assert "| 产品 | 数量 |" in text
    assert "苹果" in text and "10" in text
    assert ".xlsx" in ALLOWED_EXTS


def test_parse_xlsx_multiple_sheets_merged():
    openpyxl = pytest.importorskip("openpyxl")
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "S1"
    ws1.append(["a"]); ws1.append(["1"])
    ws2 = wb.create_sheet("S2")
    ws2.append(["b"]); ws2.append(["2"])
    buf = io.BytesIO()
    wb.save(buf)
    text = parse_upload("multi.xlsx", buf.getvalue())
    assert "## Sheet: S1" in text
    assert "## Sheet: S2" in text


def test_parse_xlsx_adds_structured_records():
    openpyxl = pytest.importorskip("openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(["item", "qty"])
    ws.append(["apple", 10])
    buf = io.BytesIO()
    wb.save(buf)

    text = parse_upload("inventory.xlsx", buf.getvalue())

    assert "\u3010\u7ed3\u6784\u5316\u8868\u683c\u8bb0\u5f55\uff1aSheet Inventory\u3011" in text
    assert "\u8bb0\u5f551: item=apple | qty=10" in text


def test_parse_xlsx_preserves_formula_only_sheets():
    openpyxl = pytest.importorskip("openpyxl")
    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "=1+2"
    summary["B1"] = "=SUM(3,4)"
    detail = wb.create_sheet("Detail")
    detail.append(["name", "qty"])
    detail.append(["apple", 2])
    buf = io.BytesIO()
    wb.save(buf)
    text = parse_upload("formula.xlsx", buf.getvalue())
    assert "## Sheet: Summary" in text
    assert "=1+2" in text
    assert "## Sheet: Detail" in text
    assert "| name | qty |" in text


def test_parse_xlsx_expands_merged_cells_and_comments():
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged"
    ws.merge_cells("A1:B1")
    ws["A1"] = "Merged Title"
    ws.append(["name", "note"])
    ws.append(["Alice", "ok"])
    ws["B3"].comment = Comment("important", "tester")
    buf = io.BytesIO()
    wb.save(buf)

    parsed = parse_upload_detail("merged.xlsx", buf.getvalue())

    assert "| Merged Title | Merged Title |" in parsed.text
    assert "批注：important" in parsed.text
    assert any("合并单元格" in warning for warning in parsed.warnings)


# ------------------------------------------------------------------ #
# Excel(.xls) —— xlwt 生成 + xlrd 解析往返
# ------------------------------------------------------------------ #
def test_parse_xls_renders_table():
    xlwt = pytest.importorskip("xlwt")
    pytest.importorskip("xlrd")
    wb = xlwt.Workbook()
    ws = wb.add_sheet("数据")
    ws.write(0, 0, "姓名"); ws.write(0, 1, "分数")
    ws.write(1, 0, "张三"); ws.write(1, 1, 88)
    buf = io.BytesIO()
    wb.save(buf)
    text = parse_upload("old.xls", buf.getvalue())
    assert "## Sheet: 数据" in text
    assert "姓名" in text and "分数" in text
    assert "张三" in text and "88" in text
    assert ".xls" in ALLOWED_EXTS


# ------------------------------------------------------------------ #
# 错误路径
# ------------------------------------------------------------------ #
def test_parse_unsupported_type():
    with pytest.raises(ValueError, match="不支持的文件类型"):
        parse_upload("program.exe", b"MZ")


def test_parse_empty_raises():
    with pytest.raises(ValueError, match="提取"):
        parse_upload("empty.py", b"   \n\t  ")


def test_allowed_exts_covers_common_dev_files():
    for ext in [".py", ".js", ".ts", ".tsx", ".java", ".html", ".css", ".go",
                ".pdf", ".docx", ".sql", ".yaml"]:
        assert ext in ALLOWED_EXTS
